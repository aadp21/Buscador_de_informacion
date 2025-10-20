# conector_sheets.py
from __future__ import annotations
import os
import json
import string
from typing import Iterable, List, Optional, Generator
import gspread
from google.oauth2.service_account import Credentials
from gspread_dataframe import set_with_dataframe
import pandas as pd
from openpyxl import load_workbook
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


# ========== Core de autenticación / cliente ==========

class GoogleSheetsClient:
    """Cliente perezoso (lazy) con cache de credenciales y acceso a APIs de gspread y Sheets v4."""
    _creds: Optional[Credentials] = None
    _gsc: Optional[gspread.Client] = None
    _svc_values = None

    def _load_creds(self) -> Credentials:
        if self._creds:
            return self._creds
        env_json = os.getenv("GOOGLE_CREDENTIALS") or os.getenv("GOOGLE_APPLICATION_CREDENTIALS_JSON")
        if env_json:
            self._creds = Credentials.from_service_account_info(json.loads(env_json), scopes=SCOPES)
        else:
            path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "config/credentials.json")
            self._creds = Credentials.from_service_account_file(path, scopes=SCOPES)
        try:
            print("➡️ Service Account:", self._creds.service_account_email)
        except Exception:
            pass
        return self._creds

    @property
    def gspread(self) -> gspread.Client:
        if not self._gsc:
            self._gsc = gspread.authorize(self._load_creds())
        return self._gsc

    @property
    def values_api(self):
        """Google Sheets API v4 values endpoint (para batch updates eficientes)."""
        if not self._svc_values:
            svc = build("sheets", "v4", credentials=self._load_creds(), cache_discovery=False)
            self._svc_values = svc.spreadsheets().values()
        return self._svc_values

    # helpers
    def open_by_key(self, sheet_id: str):
        return self.gspread.open_by_key(sheet_id)

    @staticmethod
    def a1_col(idx: int) -> str:
        # 1->A, 27->AA...
        out = ""
        while idx:
            idx, r = divmod(idx - 1, 26)
            out = string.ascii_uppercase[r] + out
        return out


client = GoogleSheetsClient()  # instancia única reutilizable


# ========== Lectores ==========

class SheetReaderBase:
    """Base para lectores de hojas."""
    def __init__(self, sheet_id: str, sheet_name: str):
        self.sheet_id = sheet_id
        self.sheet_name = sheet_name
        self._ws = None
        self._headers: List[str] = []

    @property
    def ws(self):
        if not self._ws:
            self._ws = client.open_by_key(self.sheet_id).worksheet(self.sheet_name)
        return self._ws

    def headers(self) -> List[str]:
        if self._headers:
            return self._headers
        hdr = self.ws.row_values(1) or []
        self._headers = self._dedup_headers(hdr)
        return self._headers

    @staticmethod
    def _dedup_headers(headers: List[str]) -> List[str]:
        seen = {}
        res = []
        for h in headers:
            h = h or ""
            if h in seen:
                seen[h] += 1
                res.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                res.append(h)
        return res


class FullSheetReader(SheetReaderBase):
    """Lee la hoja entera (usa get_all_values). Úsalo solo en hojas chicas."""
    def to_dataframe(self) -> pd.DataFrame:
        all_vals = self.ws.get_all_values()
        if not all_vals:
            return pd.DataFrame()
        headers = self._dedup_headers(all_vals[0])
        n = len(headers)
        rows = [(r[:n] + [""]*(n-len(r))) for r in all_vals[1:]]
        df = pd.DataFrame(rows, columns=headers)
        df.replace(["", "-"], pd.NA, inplace=True)
        return df


class PopFilteredReader(SheetReaderBase):
    """Lee SOLO filas cuyo POP == código, escaneando por chunks la columna POP."""
    def to_dataframe(self, codigo: str, chunk: int = 5000) -> pd.DataFrame:
        headers = self.headers()
        if "POP" not in [h.strip().upper() for h in headers]:
            return pd.DataFrame(columns=headers)

        # mapa de encabezados a índice 1-based
        col_map = {h.strip().upper(): i + 1 for i, h in enumerate(headers)}
        pop_col = col_map["POP"]
        last_row = self.ws.row_count
        codigo_norm = (codigo or "").strip().upper()

        matched = []
        for r0 in range(2, last_row + 1, chunk):
            r1 = min(r0 + chunk - 1, last_row)
            colL = client.a1_col(pop_col)
            # 🔧 FIX: rango RELATIVO para Worksheet.batch_get (sin prefijo de hoja)
            rng = f"{colL}{r0}:{colL}{r1}"
            blocks = self.ws.batch_get([rng])
            col = blocks[0] if blocks else []
            for i, v in enumerate(col):
                val = (v[0] if v else "").strip().upper()
                if val == codigo_norm:
                    matched.append(r0 + i)

        if not matched:
            return pd.DataFrame(columns=headers)

        # descargar esas filas completas por lotes
        last_col = client.a1_col(len(headers))
        # 🔧 FIX: rangos RELATIVOS aquí también
        ranges = [f"A{r}:{last_col}{r}" for r in matched]
        rows: List[List[str]] = []
        for i in range(0, len(ranges), 200):
            chunk_ranges = ranges[i:i+200]
            group = self.ws.batch_get(chunk_ranges)
            for fila in group:
                vals = fila[0] if fila else []
                vals = (vals[:len(headers)] + [""]*(len(headers)-len(vals)))
                rows.append(vals)

        return pd.DataFrame(rows, columns=headers)


# ========== Escritores ==========

class SheetWriterBase:
    def __init__(self, sheet_id: str, sheet_name: str):
        self.sheet_id = sheet_id
        self.sheet_name = sheet_name

    def _get_or_create_ws(self, rows=100, cols=26):
        sh = client.open_by_key(self.sheet_id)
        try:
            ws = sh.worksheet(self.sheet_name)
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(self.sheet_name, rows=rows, cols=cols)
        return ws


class StreamingWriter(SheetWriterBase):
    """Escribe en bloques usando Sheets API (memoria constante)."""
    def write_rows(self, rows_iter: Iterable[List], batch_rows: int = 2000):
        ws = self._get_or_create_ws()
        ws.clear()

        values = client.values_api
        start_row = 1
        max_cols = 0
        buf: List[List] = []

        def pad(row, cols):
            return row + [""] * (cols - len(row)) if len(row) < cols else row

        def flush():
            nonlocal start_row, buf, max_cols
            if not buf:
                return
            rect = [pad([("" if v is None else v) for v in r], max_cols) for r in buf]
            # ✅ API v4 requiere rango CALIFICADO con nombre de hoja (entre comillas)
            values.update(
                spreadsheetId=self.sheet_id,
                range=f"'{self.sheet_name}'!A{start_row}",
                valueInputOption="RAW",
                body={"values": rect},
            ).execute()
            start_row += len(rect)
            buf = []

        for row in rows_iter:
            row = list(row)
            max_cols = max(max_cols, len(row))
            buf.append(row)
            if len(buf) >= batch_rows:
                flush()
        flush()
        ws.resize(rows=start_row - 1, cols=max_cols)


class DataFrameWriter(SheetWriterBase):
    """Escritura simple con gspread (para DFs chicos)."""
    def write_df(self, df: pd.DataFrame):
        ws = self._get_or_create_ws()
        ws.clear()
        set_with_dataframe(ws, (df.copy() if df is not None else pd.DataFrame()).fillna(""))


# ========== Utilidades de Excel streaming ==========

def excel_rows_from_bytes(xio, sheet: Optional[str] = None) -> Generator[List, None, None]:
    """
    Genera filas (listas) leyendo un XLSX en modo streaming.
    La primera fila devuelta es el header.
    """
    wb = load_workbook(filename=xio, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    first = True
    ncols = None
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else v for v in row]
        if first:
            ncols = len(vals)
            yield vals
            first = False
            continue
        # normaliza a ncols
        vals = vals[:ncols] + [""] * (ncols - len(vals))
        yield vals


# ========== Wrappers retro-compatibles (tu código actual los usa) ==========

def conectar_sheets() -> gspread.Client:
    # mantenemos esta firma por compatibilidad
    return client.gspread

def leer_hoja(sheet_id: str, nombre_hoja: str) -> pd.DataFrame:
    """Versión 'full' – úsala solo en hojas pequeñas."""
    return FullSheetReader(sheet_id, nombre_hoja).to_dataframe()

def leer_filas_por_pop(sheet_id: str, nombre_hoja: str, codigo: str) -> pd.DataFrame:
    """Recomendado para hojas grandes: solo filas con POP=codigo."""
    return PopFilteredReader(sheet_id, nombre_hoja).to_dataframe(codigo)

def escribir_hoja(sheet_id: str, sheet_name: str, df: pd.DataFrame):
    DataFrameWriter(sheet_id, sheet_name).write_df(df)

def escribir_hoja_stream(sheet_id: str, sheet_name: str, rows_iter: Iterable[List], batch_rows: int = 2000):
    StreamingWriter(sheet_id, sheet_name).write_rows(rows_iter, batch_rows=batch_rows)

def escribir_excel_streaming(sheet_id: str, sheet_name: str, xio, batch_rows: int = 2000, sheet_in_xlsx: Optional[str]=None):
    """Carga un XLSX a Sheets sin DataFrame ni copias grandes."""
    rows_iter = excel_rows_from_bytes(xio, sheet=sheet_in_xlsx)
    escribir_hoja_stream(sheet_id, sheet_name, rows_iter, batch_rows=batch_rows)

