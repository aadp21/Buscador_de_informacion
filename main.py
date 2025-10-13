from fastapi import FastAPI, Request, Depends, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from conector_sheets import leer_hoja
import pandas as pd
import time
import io
import os
from typing import Dict, List
import smtplib
from email.mime.text import MIMEText
from email.utils import formatdate
from fastapi import HTTPException, status
import uuid
from typing import Any
from fastapi.responses import StreamingResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import unicodedata



SHEET_ID = "18e8Bfx5U1XLar7DOQ7PSVO5nQzluqKBHaxSOfRcreRI"
templates = Jinja2Templates(directory="templates")

app = FastAPI()

# =========================
#  Autenticación (Basic) - Multiusuario
# =========================
security = HTTPBasic()

# Fallback single admin si no hay lista
ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASS = os.getenv("ADMIN_PASS", "admin123")
UPLOAD_USERS = os.getenv(
    "UPLOAD_USERS","vrossel@olitel.cl:12345; prodriguez@olitel.cl:12345; gmunoz@olitel.cl:12345; adelgado@olitel.cl:12345"
)

def _parse_user_list(raw: str):
    users = {}
    for pair in raw.split(";"):
        pair = pair.strip()
        if not pair or ":" not in pair:
            continue
        u, p = pair.split(":", 1)
        users[u.strip()] = p.strip()
    return users

def require_auth(credentials: HTTPBasicCredentials = Depends(security)) -> str:
    """
    Devuelve el username autenticado (lo usamos para el correo).
    """
    if UPLOAD_USERS:
        users = _parse_user_list(UPLOAD_USERS)
        if credentials.username in users and credentials.password == users[credentials.username]:
            return credentials.username
    else:
        if (credentials.username == ADMIN_USER) and (credentials.password == ADMIN_PASS):
            return credentials.username

    raise HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Credenciales inválidas",
        headers={"WWW-Authenticate": "Basic"},
    )

# =========================
#  Caché (mejor rendimiento)
# =========================
CACHE_TIMEOUT = 86400  # 24 horas en segundos
data_cache: Dict[str, pd.DataFrame] = {}
last_update: Dict[str, float] = {}

def get_data(sheet_name: str) -> pd.DataFrame:
    """Devuelve DF cacheado si está fresco, si no, recarga desde Google Sheets."""
    now = time.time()
    if (
        sheet_name not in data_cache
        or sheet_name not in last_update
        or (now - last_update[sheet_name]) > CACHE_TIMEOUT
    ):
        print(f"♻️ Recargando hoja: {sheet_name}")
        df = leer_hoja(SHEET_ID, sheet_name)
        data_cache[sheet_name] = df
        last_update[sheet_name] = now
    return data_cache[sheet_name]

def invalidate_cache(sheets: List[str]):
    """Elimina entradas de caché para refrescar inmediato tras una carga."""
    for s in sheets:
        if s in data_cache:
            del data_cache[s]
        if s in last_update:
            del last_update[s]

# =========================
#  Utilidades
# =========================
# ===== Cache temporal de archivos subidos (15 min) =====
TEMP_UPLOADS: Dict[str, Dict[str, Any]] = {}
TEMP_TTL = 15 * 60  # 15 minutos

def save_temp_upload(data: bytes) -> str:
    tok = str(uuid.uuid4())
    TEMP_UPLOADS[tok] = {"data": data, "ts": time.time()}
    return tok

def get_temp_upload(tok: str) -> bytes | None:
    rec = TEMP_UPLOADS.get(tok)
    if not rec:
        return None
    if time.time() - rec["ts"] > TEMP_TTL:
        TEMP_UPLOADS.pop(tok, None)
        return None
    return rec["data"]

def purge_temp_uploads():
    now = time.time()
    for k, v in list(TEMP_UPLOADS.items()):
        if now - v["ts"] > TEMP_TTL:
            TEMP_UPLOADS.pop(k, None)

def _norm_cols_upper(cols):
    return [str(c).upper().strip() for c in cols]

def _norm_cols(cols):
    return [str(c).strip() for c in cols]

def _strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

# Modifica en filtrar_por_pop:
def filtrar_por_pop(df: pd.DataFrame, codigo: str, excluir=None):
    df = df.copy()
    # Normaliza columnas: mayúsculas + sin tildes
    df.columns = [_strip_accents(str(c).upper().strip()) for c in df.columns]
    col_pop = next((c for c in df.columns if "POP" in c), None)
    if not col_pop:
        return []
    mask = df[col_pop].astype(str).str.upper().str.strip() == _strip_accents(str(codigo).upper().strip())
    df_filtrado = df.loc[mask, df.columns]
    if excluir:
        excluir_upper = [_strip_accents(str(c).upper()) for c in excluir]
        df_filtrado = df_filtrado[[c for c in df_filtrado.columns if c not in excluir_upper]]
    return (
        df_filtrado.replace({r"\n": " "}, regex=True)
                   .fillna("")
                   .to_dict(orient="records")
    )



# =========================
#  Notificaciones por correo
# =========================
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
FROM_EMAIL = os.getenv("FROM_EMAIL", SMTP_USER or "no-reply@example.com")
NOTIFY_EMAILS = [e.strip() for e in os.getenv("NOTIFY_EMAILS", "").split(",") if e.strip()]

def send_mail(subject: str, html_body: str, to_addrs=None):
    """Envía un correo HTML a NOTIFY_EMAILS (o a to_addrs si se pasa). No revienta la request si falla."""
    if not SMTP_HOST or not (NOTIFY_EMAILS or to_addrs):
        print("ℹ️ Notificación no enviada: SMTP/DESTinatarios no configurados.")
        return
    recipients = to_addrs or NOTIFY_EMAILS
    msg = MIMEText(html_body, "html", "utf-8")
    msg["Subject"] = subject
    msg["From"] = FROM_EMAIL
    msg["To"] = ", ".join(recipients)
    msg["Date"] = formatdate(localtime=True)
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as s:
            s.starttls()
            if SMTP_USER:
                s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(FROM_EMAIL, recipients, msg.as_string())
        print(f"✉️  Mail enviado a: {recipients}")
    except smtplib.SMTPAuthenticationError as e:
        print("❌ SMTPAuthenticationError:", e)
    except Exception as e:
        print("❌ Error enviando correo:", e)


# =========================
#  Análisis de POP (vacíos/duplicados)
# =========================
def analizar_pop_df(df) -> dict:
    """
    Busca columna POP (insensible a mayúsculas y acentos), reporta vacíos y duplicados.
    Devuelve dict con: exists(bool), vacios_count, vacios_rows(list), dups_count, dups_values(list)
    """
    if df is None or df.empty:
        return {"exists": False, "vacios_count": 0, "vacios_rows": [], "dups_count": 0, "dups_values": []}

    # hallar columna POP con normalización (mayúsculas + sin tildes)
    cols_norm = {_strip_accents(str(c).upper().strip()): c for c in df.columns}
    if "POP" not in cols_norm:
        return {"exists": False, "vacios_count": 0, "vacios_rows": [], "dups_count": 0, "dups_values": []}

    col = cols_norm["POP"]
    series = df[col].astype(str)

    # vacíos (tras normalizar espacios)
    vacios_mask = series.str.strip().isin(["", "None", "nan", "NaN"])
    vacios_rows = list(df.index[vacios_mask].tolist())

    # duplicados (normalizando valor: sin tildes + mayúsculas + trim)
    norm = series.map(lambda s: _strip_accents(s)).str.upper().str.strip()
    dups_mask = norm.duplicated(keep=False) & (~vacios_mask)
    dups_values = sorted(norm[dups_mask].unique().tolist())

    return {
        "exists": True,
        "vacios_count": int(vacios_mask.sum()),
        "vacios_rows": vacios_rows[:20],   # limitar listado
        "dups_count": int(dups_mask.sum()),
        "dups_values": dups_values[:20],   # limitar listado
    }


# Intentaremos usar escribir_hoja del conector si existe
def escribir_hoja_safe(sheet_id: str, sheet_name: str, df: pd.DataFrame):
    """Escribe un DataFrame en una pestaña de Google Sheets usando conector_sheets.escribir_hoja."""
    try:
        from conector_sheets import escribir_hoja
    except Exception:
        raise RuntimeError("conector_sheets.escribir_hoja no está disponible. Implementa esta función en tu conector.")
    # Recomendado: df.fillna("") para no subir NaN
    escribir_hoja(sheet_id, sheet_name, df.fillna(""))

def backup_sheet(sheet_name: str) -> str:
    print(f"(debug) backup desactivado temporalmente para {sheet_name}")
    return ""


# =========================
#  Middleware
# =========================
@app.middleware("http")
async def log_and_block(request: Request, call_next):
    print(f"→ {request.method} {request.url.path} {request.url.query}")
    if request.method in ("HEAD", "OPTIONS"):
        return JSONResponse(content={"status": "ok"}, status_code=200)
    try:
        resp = await call_next(request)
        print(f"← {resp.status_code} {request.url.path}")
        return resp
    except Exception as e:
        print(f"✖ error en {request.url.path}: {e}")
        raise

# =========================
#  Ping
# =========================
@app.get("/")
def root():
    return {"message": "Buscador POP activo ✅"}

# =========================
#  Buscar (existente)
# =========================
@app.get("/buscar", response_class=HTMLResponse)
def buscar_pop(request: Request, codigo: str = None):
    if not codigo:
        return templates.TemplateResponse(
            "buscar.html",
            {
                "request": request,
                "codigo": None,
                "bases_result": [],
                "directorio_result": [],
                "proyecto_ranco_result": [],
                "hardware_result": [],
                "export_5g_result": [],
                "export_4g_result": [],
                "export_3g_result": [],
                "export_2g_result": [],
                "error": None,
            }
        )

    bases_result, directorio_result = [], []
    proyecto_ranco_result = []
    hardware_result, export_5g_result, export_4g_result, export_3g_result, export_2g_result = [], [], [], [], []
    error = None

    try:
        # Bases POP (solo columnas definidas)
        df_bases = get_data("Bases POP")
        df_bases.columns = [c.strip() for c in df_bases.columns]
        columnas_bases = [
            "POP","Nombre","Latitud","Longitud","Comuna","Región",
            "Tipo FDT","Tipo LLOO","ESA","Tipo","Soluc. Esp","Altura Solucion",
            "Detalle Infra ((28-12-2021))","3G1900.1","3G900","LTE3500 A/B/C",
            "NR3500","NR26000","LTE2600","LTE1900","LTE700",
            "Tecnologías Actuales Totales","TAC LTE","LAC 3G"
        ]
        # Mapa NORMALIZADO -> ORIGINAL
        cols_norm_b = {_strip_accents(c.strip().upper()): c for c in df_bases.columns}
        if "POP" in cols_norm_b:
            pop_col_b = cols_norm_b["POP"]
            mask_bases = df_bases[pop_col_b].astype(str).str.upper().str.strip() == _strip_accents(
                codigo.upper().strip())
            columnas_existentes_bases = [c for c in columnas_bases if c in df_bases.columns]
            bases_result = df_bases.loc[mask_bases, columnas_existentes_bases].fillna("").to_dict(orient="records")

        # Directorio
        df_directorio = get_data("Directorio")
        df_directorio.columns = [c.strip() for c in df_directorio.columns]

        columnas_directorio = [
            "POP", "Nombre", "Latitud", "Longitud", "Comuna", "Región",
            "Tipo FDT", "Tipo LLOO", "Tipo", "Soluc. Esp", "Detalle Infra ((28-12-2021))",
            "Tecnologías Totales Fin proyecto 2025", "CLASS 1", "CLASS 2", "CLASS 3"
        ]

        cols_norm_d = {_strip_accents(c.strip().upper()): c for c in df_directorio.columns}
        if "POP" in cols_norm_d:
            pop_col_d = cols_norm_d["POP"]
            mask_dir = df_directorio[pop_col_d].astype(str).str.upper().str.strip() == _strip_accents(
                codigo.upper().strip())
            columnas_existentes_dir = [c for c in columnas_directorio if c in df_directorio.columns]
            directorio_result = df_directorio.loc[mask_dir, columnas_existentes_dir].fillna("").to_dict(
                orient="records")

        # Proyecto_RANCO (nueva sección)
        df_ranco = get_data("Proyecto_RANCO")
        proyecto_ranco_result = filtrar_por_pop(df_ranco, codigo)

        # Base Hardware
        df_hardware = get_data("Base Hardware")
        hardware_result = filtrar_por_pop(df_hardware, codigo)

        # Export 5G
        df_5g = get_data("Export_5G")
        export_5g_result = filtrar_por_pop(df_5g, codigo, excluir=["nRSectorCarrierRef"])

        # Export 4G
        df_4g = get_data("Export_4G")
        export_4g_result = filtrar_por_pop(df_4g, codigo, excluir=["latitud", "longitud", "Región"])

        # Export 3G
        df_3g = get_data("Export_3G")
        export_3g_result = filtrar_por_pop(df_3g, codigo, excluir=["latitude", "longitude", "Región"])

        # Export 2G
        df_2g = get_data("Export_2G")
        export_2g_result = filtrar_por_pop(df_2g, codigo, excluir=["Latitude", "Longitude"])

    except Exception as e:
        error = str(e)

    return templates.TemplateResponse(
        "buscar.html",
        {
            "request": request,
            "codigo": codigo,
            "bases_result": bases_result,
            "directorio_result": directorio_result,
            "proyecto_ranco_result": proyecto_ranco_result,
            "hardware_result": hardware_result,
            "export_5g_result": export_5g_result,
            "export_4g_result": export_4g_result,
            "export_3g_result": export_3g_result,
            "export_2g_result": export_2g_result,
            "error": error,
        }
    )

# ========= Helpers Excel (una sola hoja, bloques apilados) =========
def _comparativo_rows(bases_rows: List[dict], dir_rows: List[dict]) -> List[dict]:
    campos = set()
    for r in bases_rows or []:
        campos.update(r.keys())
    for r in dir_rows or []:
        campos.update(r.keys())
    campos = [c for c in campos if c]

    def _val(rows, campo):
        if not rows:
            return ""
        vals = []
        for r in rows:
            vals.append(str(r.get(campo, "") or ""))
        return " | ".join([v for v in vals if v != ""])

    out = []
    for c in sorted(campos):
        out.append({"Campo": c, "Bases POP": _val(bases_rows, c), "Directorio": _val(dir_rows, c)})
    return out


def _rows_to_df(rows: List[dict]) -> pd.DataFrame:
    return pd.DataFrame(rows) if rows else pd.DataFrame()

def _comparativo_df(bases_rows: List[dict], dir_rows: List[dict]) -> pd.DataFrame:
    campos = set()
    for r in bases_rows or []: campos.update(r.keys())
    for r in dir_rows or []:   campos.update(r.keys())
    campos = [c for c in campos if c]

    def _val(rows, campo):
        if not rows: return ""
        vals = [str(r.get(campo, "") or "") for r in rows]
        return " | ".join([v for v in vals if v != ""])

    data = [{"Campo": c, "Bases POP": _val(bases_rows, c), "Directorio": _val(dir_rows, c)} for c in sorted(campos)]
    return pd.DataFrame(data)

def _format_sheet(ws):
    """
    Formatea la hoja:
      - Fila 1 con cabecera (negrita + fondo)
      - wrap_text en todas las celdas
      - Autoancho de columnas (estimado por contenido)
      - Altura de filas estimada según el wrap (parecido a auto-fit)
    """
    # 1) Cabecera estilizada si hay celdas
    if ws.max_row >= 1 and ws.max_column >= 1:
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 2) Habilitar wrap_text para TODO (cabecera + datos)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    # 3) Autoancho de columnas (estimado)
    #    - Igual que antes, pero lo hacemos primero para usarlo al calcular alturas
    maxw = 55  # ancho máximo para no desbordar demasiado
    minw = 10
    # Guardamos los anchos calculados para usarlos luego al estimar alturas
    computed_widths = {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = minw
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            # Heurística: ~0.9 chars por punto + 2 de margen
            width = max(width, min(maxw, len(str(v)) * 0.9 + 2))
        ws.column_dimensions[letter].width = width
        computed_widths[col] = width

    # 4) Altura de filas “auto” (estimada)
    #    - Excel no auto-ajusta filas vía openpyxl, así que calculamos un alto aproximado
    #    - Suponemos ~1.1 chars por unidad de ancho de columna
    base_height = 15          # altura de una línea
    max_height = 150          # para no generar filas gigantes
    for row in range(1, ws.max_row + 1):
        max_lines = 1
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            text = str(v)
            # líneas explícitas por saltos de línea
            hard_lines = text.split("\n")
            # ancho "útil" (aprox) de la columna en caracteres
            col_width = computed_widths.get(col, 10)
            chars_per_line = max(1, int(col_width * 1.1))  # 1.1 ≈ heurística
            est_lines = 0
            for hl in hard_lines:
                # número de líneas que ocupa ese tramo considerando el wrap
                est_lines += max(1, (len(hl) + chars_per_line - 1) // chars_per_line)
            max_lines = max(max_lines, est_lines)

        ws.row_dimensions[row].height = min(max_height, base_height * max_lines)


@app.get("/exportar_excel")
def exportar_excel(codigo: str):
    if not codigo:
        return JSONResponse({"error": "Falta parámetro codigo"}, status_code=400)

    # Bases POP (mismas columnas que en /buscar)
    df_bases = get_data("Bases POP")
    df_bases.columns = [c.strip() for c in df_bases.columns]
    columnas_bases = [
        "POP","Nombre","Latitud","Longitud","Comuna","Región",
        "Tipo FDT","Tipo LLOO","ESA","Tipo","Soluc. Esp","Altura Solucion",
        "Detalle Infra ((28-12-2021))","3G1900.1","3G900","LTE3500 A/B/C",
        "NR3500","NR26000","LTE2600","LTE1900","LTE700",
        "Tecnologías Actuales Totales","TAC LTE","LAC 3G"
    ]
    bases_rows = []
    if "POP" in df_bases.columns:
        mask = df_bases["POP"].astype(str).str.upper().str.strip() == codigo.upper().strip()
        cols_ok = [c for c in columnas_bases if c in df_bases.columns]
        bases_rows = df_bases.loc[mask, cols_ok].fillna("").to_dict(orient="records")

    # Directorio (mismas columnas que en /buscar)
    df_dir = get_data("Directorio")
    df_dir.columns = [c.strip() for c in df_dir.columns]
    columnas_dir = [
        "POP","Nombre","Latitud","Longitud","Comuna","Región",
        "Tipo FDT","Tipo LLOO","Tipo","Soluc. Esp","Detalle Infra ((28-12-2021))",
        "Tecnologías Totales Fin proyecto 2025","CLASS 1","CLASS 2","CLASS 3"
    ]
    dir_rows = []
    if "POP" in df_dir.columns:
        mask = df_dir["POP"].astype(str).str.upper().str.strip() == codigo.upper().strip()
        cols_ok = [c for c in columnas_dir if c in df_dir.columns]
        dir_rows = df_dir.loc[mask, cols_ok].fillna("").to_dict(orient="records")

    # Otras hojas (mismas exclusiones que la vista)
    proyecto_ranco_rows = filtrar_por_pop(get_data("Proyecto_RANCO"), codigo)
    hardware_rows       = filtrar_por_pop(get_data("Base Hardware"), codigo)
    export_5g_rows      = filtrar_por_pop(get_data("Export_5G"), codigo, excluir=["nRSectorCarrierRef"])
    export_4g_rows      = filtrar_por_pop(get_data("Export_4G"), codigo, excluir=["latitud", "longitud", "Región"])
    export_3g_rows      = filtrar_por_pop(get_data("Export_3G"), codigo, excluir=["latitude", "longitude", "Región"])
    export_2g_rows      = filtrar_por_pop(get_data("Export_2G"), codigo, excluir=["Latitude", "Longitude"])

    # Construir Excel multi-hoja
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) Bases POP vs Directorio (comparativo)
        df_comp = _comparativo_df(bases_rows, dir_rows)
        if df_comp.empty:
            df_comp = pd.DataFrame(columns=["Campo","Bases POP","Directorio"])
        df_comp.to_excel(writer, index=False, sheet_name="Bases POP vs Directorio")

        # 2..7) Resto de bloques
        _rows_to_df(proyecto_ranco_rows).to_excel(writer, index=False, sheet_name="Proyecto RANCO")
        _rows_to_df(hardware_rows).to_excel(writer, index=False, sheet_name="Base Hardware")
        _rows_to_df(export_5g_rows).to_excel(writer, index=False, sheet_name="Export 5G")
        _rows_to_df(export_4g_rows).to_excel(writer, index=False, sheet_name="Export 4G")
        _rows_to_df(export_3g_rows).to_excel(writer, index=False, sheet_name="Export 3G")
        _rows_to_df(export_2g_rows).to_excel(writer, index=False, sheet_name="Export 2G")

        # Formato por hoja (cabecera, wrap, auto-ancho)
        wb = writer.book
        for name in ["Bases POP vs Directorio","Proyecto RANCO","Base Hardware","Export 5G","Export 4G","Export 3G","Export 2G"]:
            ws = wb[name]
            if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
                ws["A1"] = "Sin datos para este POP"
            _format_sheet(ws)

    output.seek(0)
    filename = f"Sitio_{codigo.upper().strip()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )



# =========================
#  Interfaz de Carga
# =========================

# Página índice de Carga (protegida)
@app.get("/carga", response_class=HTMLResponse)
def carga_home(request: Request, user: str = Depends(require_auth)):
    return templates.TemplateResponse("carga.html", {"request": request})

# Subpágina de carga por tipo (form + preview + escribir)
@app.get("/carga/{tipo}", response_class=HTMLResponse)
def carga_form(request: Request, tipo: str, user: str = Depends(require_auth)):
    tipo = tipo.lower()
    if tipo not in {"bases", "directorio", "hardware", "export", "ranco"}:
        return RedirectResponse("/carga")
    return templates.TemplateResponse("carga_form.html", {"request": request, "tipo": tipo, "preview": None, "result": None, "error": None})

@app.post("/carga/{tipo}", response_class=HTMLResponse)
async def carga_upload(
        request: Request,
        tipo: str,
        file: UploadFile = File(None),  # <- opcional
        confirmar: str = Form("no"),
        token: str = Form(None),  # <- nuevo
        user: str = Depends(require_auth),
):
    tipo = tipo.lower()
    ctx = {"request": request, "tipo": tipo, "preview": None, "result": None, "error": None, "token": token}
    purge_temp_uploads()
    # 1) Preparar el archivo (xio) según si es PREVIEW o CONFIRMAR
    try:
        if confirmar != "si":
            # PREVIEW: debe venir el archivo adjunto
            if not file or not file.filename:
                ctx["error"] = "Debes subir un archivo Excel (.xlsx)."
                return templates.TemplateResponse("carga_form.html", ctx)
            fname = (file.filename or "").lower()
            if not fname.endswith(".xlsx"):
                ctx["error"] = "Debes subir un archivo Excel (.xlsx)."
                return templates.TemplateResponse("carga_form.html", ctx)
            data = await file.read()
            tok = save_temp_upload(data)
            ctx["token"] = tok
            xio = io.BytesIO(data)
        else:
            # CONFIRMAR: recuperar desde token
            if not token:
                ctx["error"] = "No se encontró token de archivo. Vuelve a subir el archivo."
                return templates.TemplateResponse("carga_form.html", ctx)
            data = get_temp_upload(token)
            if not data:
                ctx["error"] = "El archivo temporal expiró. Vuelve a subir el archivo."
                return templates.TemplateResponse("carga_form.html", ctx)
            xio = io.BytesIO(data)
    except Exception as e:
        ctx["error"] = f"Error leyendo archivo: {e}"
        return templates.TemplateResponse("carga_form.html", ctx)
    try:
        # 2) Ramas por tipo
        if tipo == "export":
            # Leer TODAS las sub-hojas
            xls: Dict[str, pd.DataFrame] = pd.read_excel(xio, sheet_name=None)
            wanted = ["Export_5G", "Export_4G", "Export_3G", "Export_2G"]
            if confirmar != "si":
                # ---- PREVIEW EXPORT ----
                preview = {}
                pops_info = {}
                for w in wanted:
                    if w in xls:
                        dfw = xls[w]
                        cols_norm = _norm_cols_upper(dfw.columns)
                        if "POP" not in cols_norm:
                            preview[w] = {
                                "ok": False, "msg": "Falta columna POP",
                                "rows": len(dfw), "cols": len(dfw.columns),
                                "columns": dfw.columns.tolist(),
                                "sample": dfw.head(5).to_dict(orient="records"),
                            }
                        else:
                            preview[w] = {
                                "ok": True, "msg": "Listo para actualizar",
                                "rows": len(dfw), "cols": len(dfw.columns),
                                "columns": dfw.columns.tolist(),
                                "sample": dfw.head(5).to_dict(orient="records"),
                            }
                            pops_info[w] = analizar_pop_df(dfw)
                    else:
                        preview[w] = {
                            "ok": False, "msg": "No está en el archivo",
                            "rows": 0, "cols": 0, "columns": [], "sample": [],
                        }
                ctx["preview"] = preview
                return templates.TemplateResponse("carga_form.html", ctx)
            # ---- CONFIRM EXPORT ----
            write_summary = {}
            touched = []
            pops_info = {}
            for w in wanted:
                if w in xls:
                    pops_info[w] = analizar_pop_df(xls[w])
                    bname = backup_sheet(w)
                    try:
                        escribir_hoja_safe(SHEET_ID, w, xls[w])
                        write_summary[
                            w] = f"Actualizado ✅ (backup: {bname if bname else 'no disponible'}) | Filas: {len(xls[w])}"
                        touched.append(w)
                    except Exception as e:
                        write_summary[w] = f"Error al escribir: {e}"
                else:
                    write_summary[w] = "Saltado: No está en el archivo"
            invalidate_cache(touched)
            ctx["result"] = write_summary
            # Email resumen
            from time import strftime, localtime
            ts = strftime("%Y-%m-%d %H:%M", localtime())
            subject = f"[Carga] Export por {user} – resultado ({ts})"
            rows_html = []
            for w in wanted:
                res = write_summary.get(w, "No procesado")
                pop_warn = pops_info.get(w)
                warn_txt = ""
                if pop_warn:
                    warn_txt = f"<br><em>POP vacíos: {pop_warn['vacios_count']}, duplicados: {pop_warn['dups_count']}</em>"
                rows_html.append(f"""
                <tr>
                  <td>{w}</td>
                  <td>{'Actualizado ✅' if 'Actualizado ✅' in res else ('Saltado' if 'Saltado' in res else 'Error')}</td>
                  <td>{res}</td>
                  <td>{warn_txt}</td>
                </tr>
                """)
            body = f"""
            <h2>Resultado de carga – Export</h2>
            <p><strong>Usuario:</strong> {user}</p>
            <p><strong>Fecha/Hora:</strong> {ts}</p>
            <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;">
              <thead style="background:#f2f2f2;">
                <tr><th>Sub-hoja</th><th>Estado</th><th>Detalle</th><th>Advertencias POP</th></tr>
              </thead>
              <tbody>
                {''.join(rows_html)}
              </tbody>
            </table>
            <p style="color:#555;">Este mensaje se envía automáticamente a: {', '.join(NOTIFY_EMAILS) if NOTIFY_EMAILS else '(destinatarios no configurados)'}.</p>
            """
            send_mail(subject, body)
            # limpiar token usado
            if token:
                TEMP_UPLOADS.pop(token, None)
            return templates.TemplateResponse("carga_form.html", ctx)
        else:
            # ====== HOJAS SIMPLES ======
            df = pd.read_excel(xio, sheet_name=0)
            cols_norm = _norm_cols_upper(df.columns)
            if confirmar != "si":
                # ---- PREVIEW SIMPLE ----
                has_pop = "POP" in cols_norm
                ctx["preview"] = {
                    "rows": len(df), "cols": len(df.columns),
                    "columns": df.columns.tolist(),
                    "sample": df.head(5).to_dict(orient="records"),
                    "ok": has_pop, "msg": "Listo para actualizar" if has_pop else "Falta columna POP",
                }
                return templates.TemplateResponse("carga_form.html", ctx)
            # ---- CONFIRM SIMPLE ----
            if "POP" not in cols_norm:
                ctx["error"] = "No se puede actualizar: falta columna POP."
                ts = time.strftime("%Y-%m-%d %H:%M", time.localtime())
                subject = f"[Carga] {tipo.capitalize()} por {user} – ERROR ({ts})"
                body = f"""
                <h2>Error en carga – {tipo.capitalize()}</h2>
                <p><strong>Usuario:</strong> {user}</p>
                <p><strong>Fecha/Hora:</strong> {ts}</p>
                <p style="color:#c62828;"><strong>Detalle:</strong> Falta columna obligatoria <code>POP</code> en la primera hoja del Excel subido.</p>
                """
                send_mail(subject, body)
                return templates.TemplateResponse("carga_form.html", ctx)
            target_map = {
                "bases": "Bases POP",
                "directorio": "Directorio",
                "hardware": "Base Hardware",
                "ranco": "Proyecto_RANCO",
            }
            target = target_map[tipo]
            pop_warn = analizar_pop_df(df)
            bname = backup_sheet(target)
            try:
                escribir_hoja_safe(SHEET_ID, target, df)
                invalidate_cache([target])
                ctx["result"] = {
                    target: f"Actualizado ✅ (backup: {bname if bname else 'no disponible'}) | Filas: {len(df)}"}
                ts = time.strftime("%Y-%m-%d %H:%M", time.localtime())
                subject = f"[Carga] {target} por {user} – ÉXITO ({ts})"
                body = f"""
                <h2>Resultado de carga – {target}</h2>
                <p><strong>Usuario:</strong> {user}</p>
                <p><strong>Fecha/Hora:</strong> {ts}</p>
                <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:14px;">
                  <tr><th style="background:#f2f2f2;">Hoja destino</th><td>{target}</td></tr>
                  <tr><th style="background:#f2f2f2;">Backup</th><td>{bname if bname else 'no disponible'}</td></tr>
                  <tr><th style="background:#f2f2f2;">Filas cargadas</th><td>{len(df)}</td></tr>
                  <tr><th style="background:#f2f2f2;">Resultado</th><td style="color:#2e7d32;"><strong>Actualizado ✅</strong></td></tr>
                </table>
                <h3 style="margin-top:18px;">Advertencias de datos (POP)</h3>
                <ul>
                  <li>POP vacíos: {pop_warn['vacios_count']}</li>
                  <li>POP duplicados: {pop_warn['dups_count']}{(' (' + ', '.join(pop_warn['dups_values']) + ')') if pop_warn['dups_values'] else ''}</li>
                </ul>
                <p style="color:#555;">Este mensaje se envía automáticamente a: {', '.join(NOTIFY_EMAILS) if NOTIFY_EMAILS else '(destinatarios no configurados)'}.</p>
                """
                send_mail(subject, body)
                if token:
                    TEMP_UPLOADS.pop(token, None)
                return templates.TemplateResponse("carga_form.html", ctx)
            except Exception as e:
                ctx["error"] = f"Error al escribir: {e}"
                ts = time.strftime("%Y-%m-%d %H:%M", time.localtime())
                subject = f"[Carga] {target} por {user} – ERROR ({ts})"
                body = f"""
                <h2>Error en carga – {target}</h2>
                <p><strong>Usuario:</strong> {user}</p>
                <p><strong>Fecha/Hora:</strong> {ts}</p>
                <p style="color:#c62828;"><strong>Detalle:</strong> {e}</p>
                """
                send_mail(subject, body)
                return templates.TemplateResponse("carga_form.html", ctx)
    except Exception as e:
        ctx["error"] = f"Error procesando archivo: {e}"
        ts = time.strftime("%Y-%m-%d %H:%M", time.localtime())
        subject = f"[Carga] {tipo.capitalize()} por {user} – ERROR ({ts})"
        body = f"""
        <h2>Error procesando archivo – {tipo.capitalize()}</h2>
        <p><strong>Usuario:</strong> {user}</p>
        <p><strong>Fecha/Hora:</strong> {ts}</p>
        <p style="color:#c62828;"><strong>Detalle:</strong> {e}</p>
        """
        send_mail(subject, body)
        return templates.TemplateResponse("carga_form.html", ctx)